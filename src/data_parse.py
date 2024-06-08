from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd

from os import listdir
import re


def get_worksheet(path: str, is_check_category: bool = False) -> Worksheet:
    """Возвращает объект странице в Excel-книге с проверкой счёта

    Получает из path файл книги и возвращает активный лист из соответствующей таблицы.
    Любые названия, не содержащие "105", "101", "21" игнорируются.

    :param path: Путь к файлу .xlsx
    :param is_check_category: Нужно ли проверять указание счета в названии таблицы?
    :return: Объект листа
    """
    if is_check_category:
        # Проверяем название таблицы, мы умеем читать только счета 21, 101, 105
        if " 105" not in path and " 21" not in path and " 101" not in path:
            raise AttributeError("Таблица неизвестного счёта: " + path)

    # Возвращаем единственный лист в книге
    workbook: Workbook = load_workbook(filename=path, read_only=True)
    return workbook.active


def get_cashflow(path: str, is_header: bool = True) -> list:
    """Возвращает данные из таблиц оборотных ведомостей в виде csv-списка. Ниже приведены заголовки.

    - Quarter: Число, получаемое как 10*Y + Q. Например, 20241 = 2024, 1 квартал
    - Category: Инвентарный/номенклатурный номер категории актива
    - Code: Код в справочнике
    - Name: Наименование актива
    - Measure: Единица измерения
    - Start Amount: Количество на первое число (01.01, 01.04, 01.07, 01.10)
    - Start Cost: Стоимость товара на первое число
    - Flow In Amount: Оборот дебет, количество
    - Flow In Cost: Оборот дебет, стоимость
    - Flow Out Amount: Оборот кредит, количество
    - Flow Out Cost: Оборот кредит, стоимость
    - End Amount: Количество на последнее число (31.03, 30.06, 30.09, 30.12)
    - End Cost: Стоимость на последнее число

    :param path: Путь к файлу вида "../data/raw/Обороты по счету/*.xlsx".
    :param  is_header: Записывать ли заголовки? При массовом получении для объединения заголовки не нужны.
    :return: Список вида [[header], [row1], [row2], ...]
    """

    sheet: Worksheet = get_worksheet(path, is_check_category=True)
    header = [
        "Quarter", "Category", "Code", "Name", "Measure", "Start Amount", "Start Cost", "Flow In Amount",
        "Flow In Cost", "Flow Out Amount", "Flow Out Cost", "End Amount", "End Cost"
    ]
    result = [header] if is_header else []

    # Таблицы имеют две различные структуры в зависимости от счета
    if " 105" in path:
        quarter_cell = sheet[1][7].value.split()  # Текущий квартал из ячейки (1,8) в виде 10*Y + Q
        quarter = int(quarter_cell[2]) + int(quarter_cell[-2]) * 10

        current_category = None
        for row in sheet.iter_rows(min_row=4, max_col=13, values_only=True):
            if row[0]:
                # Если есть ID, то это актив
                result.append([quarter, current_category] + [cell for cell in row[2:]])
            elif row[1]:
                # Если есть категория, то это категория (elif из-за строки с итогами)
                current_category = row[1].split('-')[0].strip()
    else:
        quarter_cell = sheet[2][0].value.split()  # Текущий квартал из ячейки (2,1) в виде 10*Y + Q
        quarter = int(quarter_cell[6]) + int(quarter_cell[-2]) * 10

        current_category = None
        i = 0  # Индекс изменяется нелинейно, так как данные об одном активе разбиты на четыре строки и есть категории
        m = 2 if " 21 за 1" in path else 0  # Магическое число. Таблица за 1 квартал имеет два пустых столбца
        rows = list(sheet.iter_rows(min_row=15, values_only=True))  # iter_rows намного быстрее обращения sheet[x][y]
        while i < len(rows):
            if rows[i][0] and (rows[i][0].startswith("21.") or rows[i][0].startswith("101.")):
                # Если строка начинается с "21." | "101.", то это категория
                current_category = rows[i][0].strip()
            elif rows[i][0] and i < len(rows) - 2 and rows[i + 2][4]:
                # Если есть имя, но нет номера "продукта для которого закупка", то это начало актива
                result.append([
                    quarter,  # Квартал
                    current_category,  # Категория
                    rows[i + 2][4],  # Код в справочнике
                    rows[i][0].strip(),  # Название
                    "ед",  # Единица измерения
                    (rows[i + 1][10] or 0) - (rows[i + 1][11] or 0),  # Количество на начало
                    (rows[i][10] or 0) - (rows[i][11] or 0),  # Стоимость на начало
                    rows[i + 1][12 + m],  # Количество оборота дебет
                    rows[i][12 + m],  # Стоимость оборота дебет
                    rows[i + 1][13 + m],  # Количество оборота кредит
                    rows[i][13 + m],  # Стоимость оборота кредит
                    (rows[i + 1][14 + m] or 0) - (rows[i + 1][15 + m] or 0),  # Количество на конец
                    (rows[i][14 + m] or 0) - (rows[i][15 + m] or 0)  # Стоимость на конец
                ])
                i += 3
            i += 1

    return result


def get_stocks(path: str, is_header: bool = True) -> list:
    """Возвращает данные из таблиц ведомостей складских остатков в виде csv-списка. Ниже приведены заголовки.

    - Quarter: Число, получаемое как 10*Y + Q. Например, 20241 = 2024, 1 квартал
    - Category: Инвентарный/номенклатурный номер категории актива
    - Name: Наименование актива
    - End Amount: Количество на последнее число (31.03, 30.06, 30.09, 30.12)
    - End Cost: Стоимость на последнее число

    :param path: Путь к файлу.
    :param is_header: Записывать ли заголовки? При массовом получении для объединения заголовки не нужны.
    :return: Список вида [[header], [row1], [row2], ...]
    """

    sheet: Worksheet = get_worksheet(path, is_check_category=True)
    header = ["Quarter", "Category", "Name", "End Amount", "End Cost"]
    result = [header] if is_header else []

    # Получаем квартал на основе даты в названии файла
    date_pattern = re.compile(r"\d{2}\.\d{2}\.\d{4}")
    date = date_pattern.search(path.replace('\\', '/').split('/')[-1])
    if not date:
        raise AttributeError(f"В названии таблицы {path} не указана дата")
    d, m, y = date[0].split('.')
    quarter = 10 * int(y) + int(m) // 3

    if " 105" in path:
        current_category = None
        for row in sheet.iter_rows(min_row=8, max_col=4, values_only=True):
            if row[0].startswith("105.") and not row[1]:
                # Если строка начинается со 105. и рядом нет цены, то это категория
                current_category = row[0].strip()
            elif row[1]:
                # Первый столбец всегда заполнен, если второй тоже - это актив
                result.append([quarter, current_category] + [cell for cell in row[1:4]])
    else:
        current_category = None
        for row in sheet.iter_rows(min_row=10, max_col=23, values_only=True):
            if row[0] and (row[0].startswith("21.") or row[0].startswith("101.")) and not row[2]:
                # Если строка начинается с 21. и рядом нет названия актива, то это категория
                current_category = row[0].split(',')[0].strip()
            elif row[2]:
                # Первый столбец всегда заполнен, названия активов - в третьем столбце
                result.append([quarter, current_category, row[2], row[20], row[22]])

    return result


def get_catalog(path: str, is_header: bool = True) -> list:
    """Возвращает данные из таблицы справочника СТЕ, СПГЗ, КПГЗ в виде csv-списка. Ниже приведены заголовки.

    - Name: Наименование актива
    - Params: Список наименований характеристик
    - Price: Реф. Цена
    - Category: Конечная категория справочника
    - KPGZ Code: Код КПГЗ
    - KPGZ: Категория в Классификаторе Предметов Государственного Заказа
    - SPGZ Code: Код СПГЗ
    - SPGZ: Категория в Справочнике Предметов Государственного Заказа

    :param path: Путь к файлу.
    :param is_header: Записывать ли заголовки? При массовом получении для объединения заголовки не нужны.
    :return: Список вида [[header], [row1], [row2], ...]
    """

    sheet: Worksheet = get_worksheet(path)
    header = ["Name", "Params", "Price", "Category", "KPGZ Code", "KPGZ", "SPGZ Code", "SPGZ"]
    result = [header] if is_header else []

    for row in sheet.iter_rows(min_row=2, max_col=8, values_only=True):
        if row[0]:
            result.append([row[0].strip(),  [x for x in row[1].split(';')]] + [x for x in row[2:]])

    return result


if __name__ == '__main__':
    # Устанавливают нужные таблицы для выгрузки
    is_cashflow = False
    is_stocks = False
    is_catalog = True

    # Убираем ограничение по ширине и количеству столбцов при выводе фрейма
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", None)

    # Пример получения данных оборотных ведомостей
    if is_cashflow:
        cashflow_data = []
        is_first = True
        cashflow_paths = ["../data/raw/Обороты по счету/" + x for x in listdir("../data/raw/Обороты по счету")]
        for x in cashflow_paths:
            try:
                cashflow_data += get_cashflow(x, is_header=is_first)
                is_first = False
            except PermissionError:
                print("Отказано в доступе к файлу", x)

        df_cashflow = pd.DataFrame(cashflow_data[1:], columns=cashflow_data[0])
        print(df_cashflow)

        df_cashflow.to_csv("../data/processed/cashflow.csv", index=False)

    # Пример получения данных ведомостей складских остатков
    if is_stocks:
        stocks_data = []
        is_first = True
        stocks_paths = ["../data/raw/Складские остатки/" + x for x in listdir("../data/raw/Складские остатки")]
        for x in stocks_paths:
            try:
                stocks_data += get_stocks(x, is_header=is_first)
                is_first = False
            except PermissionError:
                print("Отказано в доступе к файлу", x)

        df_stocks = pd.DataFrame(stocks_data[1:], columns=stocks_data[0])
        print(df_stocks)

        df_stocks.to_csv("../data/processed/stocks.csv", index=False)

    # Пример получения данных справочника КПГЗ
    if is_catalog:
        catalog_path = "../data/raw/КПГЗ, СПГЗ, СТЕ.xlsx"
        catalog_data = get_catalog(catalog_path)

        df_catalog = pd.DataFrame(catalog_data[1:], columns=catalog_data[0])
        print(df_catalog)

        df_catalog.to_csv("../data/processed/catalog.csv", index=False)

